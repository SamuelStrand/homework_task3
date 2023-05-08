import json

from django.contrib.auth import get_user_model

from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import update_last_login, User
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpResponse, JsonResponse, HttpRequest
from django.shortcuts import render, redirect
from django.urls import reverse
from openpyxl.workbook import Workbook

from django_app import models

from django_app.models import Post


def list_compr(request: HttpRequest) -> JsonResponse:
    data = [
        {
            'id': x,
            'name': f'Emil {x}',
            'age': x
        }
        for x
        in range(1, 100)]

    with open('temp/logs.txt', 'w') as logs_file:
        json.dump(data, logs_file)

    return JsonResponse(data=data, safe=False)


def login_f(request: HttpRequest) -> HttpResponse:
    if request.method == 'GET':
        context = {}
        return render(request, 'login.html', context=context)
    if request.method == 'POST':
        username = request.POST.get('username', "")
        password = request.POST.get('password', "")
        if username and password:
            user_obj = authenticate(username=username, password=password)
            if user_obj:
                login(request, user_obj)
                update_last_login(sender=None, user=user_obj)
                return redirect(reverse('django_app:post_list', args=()))
            else:
                raise Exception('данные не совпадают')
        else:
            raise Exception('no data')


def logout_f(request: HttpRequest) -> HttpResponse:
    logout(request)
    return redirect(reverse('django_app:login', args=()))


def post_list(request: HttpRequest) -> HttpResponse:
    posts = models.Post.objects.all()

    if request.method == 'POST':
        search_by_title = request.POST.get('search', None)
        if search_by_title is not None:
            posts = posts.filter(title__contains=str(search_by_title))
        filter_by_user = request.POST.get('filter', None)
        if filter_by_user is not None:
            posts = posts.filter(user=User.objects.get(username=filter_by_user))

    selected_page_number = request.GET.get('page', 1)
    selected_limit_objects_per_page = request.GET.get('limit', 6)

    page = CustomPaginator.paginate(
        object_list=posts, per_page=selected_limit_objects_per_page, page_number=selected_page_number
    )

    users = User.objects.all()
    context = {'page': page, 'username': request.user, 'users': users}
    return render(request, 'post_list.html', context=context)


def register(request: HttpRequest) -> HttpResponse:
    username = request.POST.get('username', '')
    password = request.POST.get('password', '')
    if username and password:
        User.objects.create(
            username=username,
            password=make_password(password)
        )
        return redirect(reverse('django_app:login'))
    return render(request, 'register.html')


def home(request: HttpRequest) -> HttpResponse:
    context = {}
    return render(request, 'components/base.html', context=context)


def post_detail(request: HttpRequest, pk: int):
    post = models.Post.objects.get(id=pk)
    context = {'post': post}
    return render(request, 'post_detail.html', context=context)


def post_create(request: HttpRequest) -> HttpResponse:
    if request.method == 'GET':
        context = {}
        return render(request, 'post_create.html', context=context)
    elif request.method == 'POST':
        print('printed', request.POST)
        title = request.POST.get('title', None)
        description = request.POST.get('description', "")
        models.Post.objects.create(
            user=request.user,
            title=title,
            description=description,
        )

        context = {}
        return redirect(reverse('django_app:post_list', args=()))


def post_delete(request: HttpRequest, pk: int) -> HttpResponse:
    models.Post.objects.get(id=pk).delete()
    return redirect(reverse('django_app:post_list', args=()))


def get_data_excel(request):
    user_workbook = Workbook()
    user_worksheet = user_workbook.active

    post_workbook = Workbook()
    post_worksheet = post_workbook.active

    users = User.objects.all()
    posts = Post.objects.all()

    for user_header_index, user_header_value in enumerate(['user_id', 'username', 'password', 'email', 'is_staff'], 1):
        user_worksheet.cell(1, user_header_index, user_header_value)
    for row_index, user in enumerate(users, 2):
        user_id = user.id
        username = user.username
        password = user.password
        email = user.email
        is_staff = user.is_staff
        cols = [user_id, username, password, email, is_staff]
        for col_index, value in enumerate(cols, 1):
            user_worksheet.cell(row_index, col_index, value)
    user_workbook.save('temp/users_data.xlsx')

    for post_header_index, post_header_value in enumerate(['post_user', 'title', 'description'], 1):
        post_worksheet.cell(1, post_header_index, post_header_value)
    for row_index_posts, post in enumerate(posts, 2):
        post_user = str(post.user)
        title = post.title
        description = post.description
        post_cols = [post_user, title, description]
        for col_index_posts, post_value in enumerate(post_cols, 1):
            post_cell = post_worksheet.cell(row_index_posts, col_index_posts, post_value)
    post_workbook.save('temp/posts_data.xlsx')
    return HttpResponse('<h1>Данные успешно выгружены в файл!</h1>')


class CustomPaginator:
    @staticmethod
    def paginate(object_list: any, per_page=5, page_number=1):
        # https://docs.djangoproject.com/en/4.1/topics/pagination/
        paginator_instance = Paginator(object_list=object_list, per_page=per_page)
        try:
            page = paginator_instance.page(number=page_number)
        except PageNotAnInteger:
            page = paginator_instance.page(number=1)
        except EmptyPage:
            page = paginator_instance.page(number=paginator_instance.num_pages)
        return page


def all_users(request: HttpRequest) -> HttpResponse:
    user = get_user_model()
    users = user.objects.all()
    context = {'users': users}
    return render(request, 'all_users.html', context=context)
